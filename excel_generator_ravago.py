import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
from openpyxl.utils import column_index_from_string as colidx
from openpyxl.worksheet.page import PageMargins

# =============================
# Utilidades de formato
# =============================

# Meses en español para la fecha
MESES_ES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
]

def fecha_es(dt: datetime) -> str:
    """Devuelve 'dd de <mes> de yyyy' en español."""
    return f"{dt.day:02d} de {MESES_ES[dt.month-1]} de {dt.year}"

def style_cell(cell, text, font, alignment, fill=None, border=None):
    """Aplica estilos completos a una celda."""
    cell.value = text
    cell.font = font
    cell.alignment = alignment
    if fill is not None:
        cell.fill = fill
    if border is not None:
        cell.border = border

def _letters(s: str) -> str:
    return "".join(ch for ch in s if ch.isalpha())

def _digits(s: str) -> int:
    return int("".join(ch for ch in s if ch.isdigit()))

def draw_outer_frame(ws, tl: str, br: str):
    """
    Dibuja un marco negro de grosor 'medium' alrededor del rectángulo
    delimitado por las celdas tl (top-left) y br (bottom-right),
    combinando con cualquier borde ya existente para evitar 'cortes'.
    """
    top = _digits(tl)
    bottom = _digits(br)
    left = colidx(_letters(tl))
    right = colidx(_letters(br))

    side = Side(style="medium", color="000000")

    # Fila superior (borde superior)
    for c in range(left, right + 1):
        cell = ws.cell(row=top, column=c)
        prev = cell.border
        cell.border = Border(
            left=prev.left, right=prev.right, top=side, bottom=prev.bottom
        )

    # Fila inferior (borde inferior)
    for c in range(left, right + 1):
        cell = ws.cell(row=bottom, column=c)
        prev = cell.border
        cell.border = Border(
            left=prev.left, right=prev.right, top=prev.top, bottom=side
        )

    # Columna izquierda (borde izquierdo)
    for r in range(top, bottom + 1):
        cell = ws.cell(row=r, column=left)
        prev = cell.border
        cell.border = Border(
            left=side, right=prev.right, top=prev.top, bottom=prev.bottom
        )

    # Columna derecha (borde derecho)
    for r in range(top, bottom + 1):
        cell = ws.cell(row=r, column=right)
        prev = cell.border
        cell.border = Border(
            left=prev.left, right=side, top=prev.top, bottom=prev.bottom
        )

# =============================
# Generador del reporte
# =============================

def find_column(df, possible_names):
    """Busca una columna en el DataFrame usando una lista de posibles nombres."""
    for col_name in possible_names:
        for actual_col in df.columns:
            if col_name.upper() in actual_col.upper():
                return actual_col
    return None

def get_document_count(df):
    """Obtiene el número de documentos únicos."""
    possible_names = ['NO. CASO', 'NUMERO CASO', 'CASO', 'ID', 'NUMERO', 'DOCUMENTO']
    col_name = find_column(df, possible_names)
    
    if col_name:
        return df[col_name].nunique()
    else:
        return len(df)

def create_ravago_report(data: pd.DataFrame, anio: int, mes: str, funcionarios: dict | None = None):
    """
    Genera un Excel con dos hojas:
      - 'Facturación' con el layout exacto solicitado
      - 'Anexo 1' con el layout exacto solicitado
    """
    wb = Workbook()

    # -----------------
    # Estilos globales
    # -----------------
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    total_font  = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    data_font   = Font(name='Calibri', size=11, bold=False, color="000000")
    title_font  = Font(name='Calibri', size=12, bold=True, color="000000")
    info_font   = Font(name='Calibri', size=11)
    italic_font = Font(name='Calibri', size=9, italic=True, color="000000")
    # Azul corporativo #002060 en ARGB (FF = opaco)
    HEADER_HEX = "FF002060"
    header_fill = PatternFill(fill_type="solid", start_color=HEADER_HEX, end_color=HEADER_HEX)
    total_fill  = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    white_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    center = Alignment(horizontal='center',  vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',    vertical='center', wrap_text=False)
    right  = Alignment(horizontal='right',   vertical='center', wrap_text=False)

    # -----------------
    # Datos auxiliares
    # -----------------
    rep_name = (funcionarios or {}).get("reporta", "________________")
    rev_name = (funcionarios or {}).get("revisor", "________________")
    fecha_dt = (funcionarios or {}).get("fecha", datetime.now())

    # Contadores/columnas
    try:
        num_docs = get_document_count(data)
    except Exception:
        num_docs = len(data)

    if 'VALOR' in data.columns:
        total_valor = float(data['VALOR'].fillna(0).sum())
        valor_col = 'VALOR'
    else:
        valor_col = find_column(data, ['VALOR', 'TOTAL', 'IMPORTE', 'MONTO'])
        total_valor = float(data.get(valor_col, pd.Series(dtype=float)).fillna(0).sum()) if valor_col else 0.0

    nombre_col = find_column(data, ['NOMBRE', 'NOMBRE CONTRAPARTE', 'CLIENTE'])
    tipo_doc_col = find_column(data, ['TIPO DE DOCUMENTO', 'TIPO DOCUMENTO', 'DOCUMENTO'])

    # =========================
    # Hoja 1: Facturación
    # =========================
    ws1 = wb.active
    ws1.title = "Facturación"
    ws1.sheet_view.showGridLines = False

    # Anchos y márgenes internos (G como margen derecho)
    ws1.column_dimensions['A'].width = 2
    ws1.column_dimensions['B'].width = 4
    ws1.column_dimensions['C'].width = 38
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 30
    ws1.column_dimensions['F'].width = 16
    ws1.column_dimensions['G'].width = 8

    ws1.page_margins = PageMargins(left=0.6, right=0.6, top=0.6, bottom=0.6, header=0.3, footer=0.3)
    ws1.print_options.horizontalCentered = True

    # Logo
    try:
        img = Image('assets/biu_logo.png')
        img.width, img.height = 100, 58
        ws1.add_image(img, 'F3')
    except Exception:
        style_cell(ws1['F3'], "BIU", title_font, center)

    # Cabecera exacta (C3, C4, C5)
    style_cell(ws1['C3'], f"Fecha de corte del reporte: {fecha_es(fecha_dt)}", info_font, left)
    style_cell(ws1['C4'], f"Funcionario que reporta: {rep_name}", info_font, left)
    style_cell(ws1['C5'], f"Funcionario revisor: {rev_name}", info_font, left)

    # Tabla superior (fila 8: headers; 9: valores; 10: total por facturar)
    style_cell(ws1['C8'], 'Año', header_font, center, header_fill, thin_border)
    style_cell(ws1['D8'], 'Mes', header_font, center, header_fill, thin_border)
    style_cell(ws1['E8'], 'Documentos Revisados (Ver Anexo 1)', header_font, center, header_fill, thin_border)

    style_cell(ws1['C9'], anio, data_font, center, white_fill, thin_border)
    style_cell(ws1['D9'], mes, data_font, center, white_fill, thin_border)
    style_cell(ws1['E9'], num_docs, data_font, center, white_fill, thin_border)

    # Fila 10: D10 y E10; C10 SIN bordes ni relleno (blanco)
    style_cell(ws1['D10'], 'Total Por Facturar', total_font, center, total_fill, thin_border)
    style_cell(ws1['E10'], num_docs, total_font, center, total_fill, thin_border)
    style_cell(ws1['C10'], '', data_font, left, white_fill, None)  # <-- sin bordes

    # Segunda tabla (filas 12-14)
    ws1.merge_cells('C12:D12')
    style_cell(ws1['C12'], 'Concepto', header_font, center, header_fill, thin_border)
    style_cell(ws1['E12'], 'Total (antes de I.V.A)', header_font, center, header_fill, thin_border)

    ws1.merge_cells('C13:D13')
    style_cell(
        ws1['C13'],
        f"Revisión de {num_docs} documentos durante el mes de {mes} de {anio}",
        data_font, Alignment(horizontal='left', vertical='center', wrap_text=True),
        white_fill, thin_border
    )
    c = ws1['E13']
    style_cell(c, total_valor, data_font, right, white_fill, thin_border)
    c.number_format = '"USD" #,##0'

    # Fila 14: C14 SIN bordes; D14 SUBTOTAL; E14 valor
    style_cell(ws1['C14'], '', data_font, left, white_fill, None)  # <-- sin bordes
    style_cell(ws1['D14'], 'SUBTOTAL', total_font, right, total_fill, thin_border)
    c = ws1['E14']
    style_cell(c, total_valor, total_font, right, total_fill, thin_border)
    c.number_format = '"USD" #,##0'

    # Notas
    ws1.merge_cells('C16:D16')
    style_cell(ws1['C16'], "TRM Aplicable: Según la propuesta, es aquella de emisión de la factura.",
               info_font, Alignment(horizontal='left', vertical='center', wrap_text=True))

    ws1.merge_cells('C18:E20')
    footer_text = ("biu usually issues monthly invoices for the provision of the Services; "
                   "the amounts indicated in US dollars shall be converted based on the official "
                   "prevailing market rate as of the date of issuance of the invoice.")
    style_cell(ws1['C18'], footer_text, italic_font,
               Alignment(horizontal='left', vertical='top', wrap_text=True))

    # Marco exterior (con G como margen derecho)
    draw_outer_frame(ws1, "B2", "G21")

    # =========================
    # Hoja 2: Anexo 1
    # =========================
    ws2 = wb.create_sheet(title="Anexo 1")
    ws2.sheet_view.showGridLines = False

    ws2.column_dimensions['A'].width = 2
    ws2.column_dimensions['B'].width = 4
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 36
    ws2.column_dimensions['E'].width = 44
    ws2.column_dimensions['F'].width = 16
    ws2.column_dimensions['G'].width = 8

    ws2.page_margins = PageMargins(left=0.6, right=0.6, top=0.6, bottom=0.6, header=0.3, footer=0.3)
    ws2.print_options.horizontalCentered = True

    try:
        img2 = Image('assets/biu_logo.png')
        img2.width, img2.height = 100, 58
        ws2.add_image(img2, 'F3')
    except Exception:
        style_cell(ws2['F3'], "BIU", title_font, center)

    style_cell(ws2['C3'], f"Fecha de corte del reporte: {fecha_es(fecha_dt)}", info_font, left)
    style_cell(ws2['C4'], f"Funcionario que reporta: {rep_name}", info_font, left)
    style_cell(ws2['C5'], f"Funcionario revisor: {rev_name}", info_font, left)

    ws2.merge_cells('C6:F6')
    style_cell(ws2['C6'], "HONORARIOS", title_font, center)

    style_cell(ws2['C8'], 'FECHA', header_font, center, header_fill, thin_border)
    style_cell(ws2['D8'], 'NOMBRE CONTRAPARTE', header_font, center, header_fill, thin_border)
    style_cell(ws2['E8'], 'TIPO DE DOCUMENTO', header_font, center, header_fill, thin_border)
    style_cell(ws2['F8'], 'TOTAL', header_font, center, header_fill, thin_border)

    start_row = 9
    r = start_row
    # Filas de detalle
    for idx, row_data in data.iterrows():
        # FECHA: consecutivo 1,2,3,...
        style_cell(ws2.cell(row=r, column=3), r - start_row + 1, data_font, center, white_fill, thin_border)

        nombre_value = row_data.get(nombre_col, '') if nombre_col else ''
        style_cell(ws2.cell(row=r, column=4), nombre_value, data_font,
                   Alignment(horizontal='left', vertical='center', wrap_text=True),
                   white_fill, thin_border)

        tipo_doc_value = row_data.get(tipo_doc_col, '') if tipo_doc_col else ''
        style_cell(ws2.cell(row=r, column=5), tipo_doc_value, data_font,
                   Alignment(horizontal='left', vertical='center', wrap_text=True),
                   white_fill, thin_border)

        valor_value = float(row_data.get(valor_col, 0)) if valor_col else 0.0
        c_val = ws2.cell(row=r, column=6)
        style_cell(c_val, valor_value, data_font, right, white_fill, thin_border)
        c_val.number_format = '"USD" #,##0'
        r += 1

    # Fila de SUBTOTAL
    subtotal_row = r

    # C11 y D11 BLANCAS y SIN BORDES (tal como pediste)
    # OJO: es la fila 'subtotal_row' para cualquier cantidad de filas
    style_cell(ws2.cell(row=subtotal_row, column=3), '', data_font, left, white_fill, None)  # C#
    style_cell(ws2.cell(row=subtotal_row, column=4), '', data_font, left, white_fill, None)  # D#

    # SUBTOTAL en E# y total en F# con gris y bordes finos
    style_cell(ws2.cell(row=subtotal_row, column=5), 'SUBTOTAL', total_font, right, total_fill, thin_border)
    c_sub = ws2.cell(row=subtotal_row, column=6)
    style_cell(c_sub, total_valor, total_font, right, total_fill, thin_border)
    c_sub.number_format = '"USD" #,##0'

    # Marco exterior (hasta G y dejando una fila extra bajo el subtotal)
    bottom_row_for_frame = subtotal_row + 1
    draw_outer_frame(ws2, "B2", f"G{bottom_row_for_frame}")

    # =========================
    # Guardar en memoria
    # =========================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
