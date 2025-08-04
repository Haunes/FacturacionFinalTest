from openpyxl.drawing.image import Image
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import column_index_from_string as colidx
from openpyxl.styles import Side, Border
from .excel_styles import ExcelStyleManager

# Meses en español
MESES_ES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
]

def fecha_es(dt) -> str:
    """Devuelve 'dd de <mes> de yyyy' en español."""
    return f"{dt.day:02d} de {MESES_ES[dt.month-1]} de {dt.year}"

class ExcelSheetBuilder:
    """Construye las hojas de Excel para reportes de Ravago."""
    
    def __init__(self):
        self.style_manager = ExcelStyleManager()
    
    def build_facturacion_sheet(self, ws, report_data):
        """Construye la hoja de Facturación."""
        self._setup_sheet_layout(ws)
        self._add_logo(ws, 'F3')
        self._add_header_info(ws, report_data)
        self._add_summary_tables(ws, report_data)
        self._add_notes(ws, report_data)
        self._add_outer_frame(ws, "B2", "G21")
    
    def build_anexo_sheet(self, ws, report_data, data):
        """Construye la hoja de Anexo 1."""
        self._setup_anexo_layout(ws)
        self._add_logo(ws, 'F3')
        self._add_header_info(ws, report_data)
        self._add_anexo_title(ws)
        self._add_detail_table(ws, report_data, data)
    
    def _setup_sheet_layout(self, ws):
        """Configura el layout básico de la hoja de Facturación."""
        ws.sheet_view.showGridLines = False
        
        # Anchos de columna
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 4
        ws.column_dimensions['C'].width = 38
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 8
        
        # Márgenes
        ws.page_margins = PageMargins(left=0.6, right=0.6, top=0.6, bottom=0.6, header=0.3, footer=0.3)
        ws.print_options.horizontalCentered = True
    
    def _setup_anexo_layout(self, ws):
        """Configura el layout básico de la hoja de Anexo."""
        ws.sheet_view.showGridLines = False
        
        # Anchos de columna para Anexo
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 4
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 36
        ws.column_dimensions['E'].width = 44
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 8
        
        # Márgenes
        ws.page_margins = PageMargins(left=0.6, right=0.6, top=0.6, bottom=0.6, header=0.3, footer=0.3)
        ws.print_options.horizontalCentered = True
    
    def _add_logo(self, ws, cell_position):
        """Agrega el logo BIU a la hoja."""
        try:
            img = Image('assets/biu_logo.png')
            img.width, img.height = 100, 58
            ws.add_image(img, cell_position)
        except Exception:
            self.style_manager.style_cell(
                ws[cell_position], "BIU", 
                self.style_manager.title_font, 
                self.style_manager.center
            )
    
    def _add_header_info(self, ws, report_data):
        """Agrega la información del encabezado."""
        fecha_str = fecha_es(report_data['fecha_dt'])
        
        self.style_manager.style_cell(
            ws['C3'], f"Fecha de corte del reporte: {fecha_str}", 
            self.style_manager.info_font, self.style_manager.left
        )
        self.style_manager.style_cell(
            ws['C4'], f"Funcionario que reporta: {report_data['rep_name']}", 
            self.style_manager.info_font, self.style_manager.left
        )
        self.style_manager.style_cell(
            ws['C5'], f"Funcionario revisor: {report_data['rev_name']}", 
            self.style_manager.info_font, self.style_manager.left
        )
    
    def _add_summary_tables(self, ws, report_data):
        """Agrega las tablas de resumen a la hoja de Facturación."""
        # Primera tabla (fila 8-10)
        self._add_first_summary_table(ws, report_data)
        
        # Segunda tabla (fila 12-14)
        self._add_second_summary_table(ws, report_data)
    
    def _add_first_summary_table(self, ws, report_data):
        """Agrega la primera tabla de resumen."""
        # Headers
        self.style_manager.style_cell(
            ws['C8'], 'Año', self.style_manager.header_font, 
            self.style_manager.center, self.style_manager.header_fill, 
            self.style_manager.thin_border
        )
        self.style_manager.style_cell(
            ws['D8'], 'Mes', self.style_manager.header_font, 
            self.style_manager.center, self.style_manager.header_fill, 
            self.style_manager.thin_border
        )
        self.style_manager.style_cell(
            ws['E8'], 'Documentos Revisados (Ver Anexo 1)', self.style_manager.header_font, 
            self.style_manager.center, self.style_manager.header_fill, 
            self.style_manager.thin_border
        )
        
        # Valores
        self.style_manager.style_cell(
            ws['C9'], report_data['anio'], self.style_manager.data_font, 
            self.style_manager.center, self.style_manager.white_fill, 
            self.style_manager.thin_border
        )
        self.style_manager.style_cell(
            ws['D9'], report_data['mes'], self.style_manager.data_font, 
            self.style_manager.center, self.style_manager.white_fill, 
            self.style_manager.thin_border
        )
        self.style_manager.style_cell(
            ws['E9'], report_data['num_docs'], self.style_manager.data_font, 
            self.style_manager.center, self.style_manager.white_fill, 
            self.style_manager.thin_border
        )
        
        # Fila total
        self.style_manager.style_cell(
            ws['C10'], '', self.style_manager.data_font, 
            self.style_manager.left, self.style_manager.white_fill, None
        )
        self.style_manager.style_cell(
            ws['D10'], 'Total Por Facturar', self.style_manager.total_font, 
            self.style_manager.center, self.style_manager.total_fill, 
            self.style_manager.thin_border
        )
        self.style_manager.style_cell(
            ws['E10'], report_data['num_docs'], self.style_manager.total_font, 
            self.style_manager.center, self.style_manager.total_fill, 
            self.style_manager.thin_border
        )
    
    def _add_second_summary_table(self, ws, report_data):
        """Agrega la segunda tabla de resumen."""
        # Fusionar celdas para el header
        ws.merge_cells('C12:D12')
        self.style_manager.style_cell(
            ws['C12'], 'Concepto', self.style_manager.header_font, 
            self.style_manager.center, self.style_manager.header_fill, 
            self.style_manager.thin_border
        )
        self.style_manager.style_cell(
            ws['E12'], 'Total (antes de I.V.A)', self.style_manager.header_font, 
            self.style_manager.center, self.style_manager.header_fill, 
            self.style_manager.thin_border
        )
        
        # Contenido
        ws.merge_cells('C13:D13')
        concepto_text = f"Revisión de {report_data['num_docs']} documentos durante el mes de {report_data['mes']} de {report_data['anio']}"
        self.style_manager.style_cell(
            ws['C13'], concepto_text, self.style_manager.data_font, 
            self.style_manager.left, self.style_manager.white_fill, 
            self.style_manager.thin_border
        )
        
        c = ws['E13']
        self.style_manager.style_cell(
            c, report_data['total_valor'], self.style_manager.data_font, 
            self.style_manager.right, self.style_manager.white_fill, 
            self.style_manager.thin_border
        )
        c.number_format = '"USD" #,##0'
        
        # Subtotal
        self.style_manager.style_cell(
            ws['C14'], '', self.style_manager.data_font, 
            self.style_manager.left, self.style_manager.white_fill, None
        )
        self.style_manager.style_cell(
            ws['D14'], 'SUBTOTAL', self.style_manager.total_font, 
            self.style_manager.right, self.style_manager.total_fill, 
            self.style_manager.thin_border
        )
        
        c = ws['E14']
        self.style_manager.style_cell(
            c, report_data['total_valor'], self.style_manager.total_font, 
            self.style_manager.right, self.style_manager.total_fill, 
            self.style_manager.thin_border
        )
        c.number_format = '"USD" #,##0'
    
    def _add_notes(self, ws, report_data):
        """Agrega las notas al pie de la hoja."""
        ws.merge_cells('C16:D16')
        self.style_manager.style_cell(
            ws['C16'], "TRM Aplicable: Según la propuesta, es aquella de emisión de la factura.",
            self.style_manager.info_font, self.style_manager.left
        )
        
        ws.merge_cells('C18:E20')
        footer_text = ("biu usually issues monthly invoices for the provision of the Services; "
                      "the amounts indicated in US dollars shall be converted based on the official "
                      "prevailing market rate as of the date of issuance of the invoice.")
        self.style_manager.style_cell(
            ws['C18'], footer_text, self.style_manager.italic_font, 
            self.style_manager.left
        )
    
    def _add_anexo_title(self, ws):
        """Agrega el título de la hoja Anexo."""
        ws.merge_cells('C6:F6')
        self.style_manager.style_cell(
            ws['C6'], "HONORARIOS", self.style_manager.title_font, 
            self.style_manager.center
        )
    
    def _add_detail_table(self, ws, report_data, data):
        """Agrega la tabla de detalle en la hoja Anexo."""
        # Headers
        self.style_manager.style_cell(
            ws['C8'], 'FECHA', self.style_manager.header_font, 
            self.style_manager.center, self.style_manager.header_fill, 
            self.style_manager.thin_border
        )
        self.style_manager.style_cell(
            ws['D8'], 'NOMBRE CONTRAPARTE', self.style_manager.header_font, 
            self.style_manager.center, self.style_manager.header_fill, 
            self.style_manager.thin_border
        )
        self.style_manager.style_cell(
            ws['E8'], 'TIPO DE DOCUMENTO', self.style_manager.header_font, 
            self.style_manager.center, self.style_manager.header_fill, 
            self.style_manager.thin_border
        )
        self.style_manager.style_cell(
            ws['F8'], 'TOTAL', self.style_manager.header_font, 
            self.style_manager.center, self.style_manager.header_fill, 
            self.style_manager.thin_border
        )
        
        # Filas de detalle
        start_row = 9
        r = start_row
        for idx, row_data in data.iterrows():
            # FECHA: consecutivo 1,2,3,...
            self.style_manager.style_cell(
                ws.cell(row=r, column=3), r - start_row + 1, 
                self.style_manager.data_font, self.style_manager.center, 
                self.style_manager.white_fill, self.style_manager.thin_border
            )
            
            nombre_value = row_data.get(report_data['nombre_col'], '') if report_data['nombre_col'] else ''
            self.style_manager.style_cell(
                ws.cell(row=r, column=4), nombre_value, 
                self.style_manager.data_font, self.style_manager.left, 
                self.style_manager.white_fill, self.style_manager.thin_border
            )
            
            tipo_doc_value = row_data.get(report_data['tipo_doc_col'], '') if report_data['tipo_doc_col'] else ''
            self.style_manager.style_cell(
                ws.cell(row=r, column=5), tipo_doc_value, 
                self.style_manager.data_font, self.style_manager.left, 
                self.style_manager.white_fill, self.style_manager.thin_border
            )
            
            valor_value = float(row_data.get(report_data['valor_col'], 0)) if report_data['valor_col'] else 0.0
            c_val = ws.cell(row=r, column=6)
            self.style_manager.style_cell(
                c_val, valor_value, self.style_manager.data_font, 
                self.style_manager.right, self.style_manager.white_fill, 
                self.style_manager.thin_border
            )
            c_val.number_format = '"USD" #,##0'
            r += 1
        
        # Fila de SUBTOTAL
        subtotal_row = r
        
        # Celdas blancas sin bordes
        self.style_manager.style_cell(
            ws.cell(row=subtotal_row, column=3), '', 
            self.style_manager.data_font, self.style_manager.left, 
            self.style_manager.white_fill, None
        )
        self.style_manager.style_cell(
            ws.cell(row=subtotal_row, column=4), '', 
            self.style_manager.data_font, self.style_manager.left, 
            self.style_manager.white_fill, None
        )
        
        # SUBTOTAL
        self.style_manager.style_cell(
            ws.cell(row=subtotal_row, column=5), 'SUBTOTAL', 
            self.style_manager.total_font, self.style_manager.right, 
            self.style_manager.total_fill, self.style_manager.thin_border
        )
        
        c_sub = ws.cell(row=subtotal_row, column=6)
        self.style_manager.style_cell(
            c_sub, report_data['total_valor'], 
            self.style_manager.total_font, self.style_manager.right, 
            self.style_manager.total_fill, self.style_manager.thin_border
        )
        c_sub.number_format = '"USD" #,##0'
        
        # Marco exterior
        bottom_row_for_frame = subtotal_row + 1
        self._add_outer_frame(ws, "B2", f"G{bottom_row_for_frame}")
    
    def _add_outer_frame(self, ws, tl: str, br: str):
        """Dibuja un marco exterior alrededor del área especificada."""
        def _letters(s: str) -> str:
            return "".join(ch for ch in s if ch.isalpha())
        
        def _digits(s: str) -> int:
            return int("".join(ch for ch in s if ch.isdigit()))
        
        top = _digits(tl)
        bottom = _digits(br)
        left = colidx(_letters(tl))
        right = colidx(_letters(br))
        
        side = Side(style="medium", color="000000")
        
        # Bordes superiores e inferiores
        for c in range(left, right + 1):
            # Borde superior
            cell = ws.cell(row=top, column=c)
            prev = cell.border
            cell.border = Border(
                left=prev.left, right=prev.right, top=side, bottom=prev.bottom
            )
            
            # Borde inferior
            cell = ws.cell(row=bottom, column=c)
            prev = cell.border
            cell.border = Border(
                left=prev.left, right=prev.right, top=prev.top, bottom=side
            )
        
        # Bordes izquierdo y derecho
        for r in range(top, bottom + 1):
            # Borde izquierdo
            cell = ws.cell(row=r, column=left)
            prev = cell.border
            cell.border = Border(
                left=side, right=prev.right, top=prev.top, bottom=prev.bottom
            )
            
            # Borde derecho
            cell = ws.cell(row=r, column=right)
            prev = cell.border
            cell.border = Border(
                left=prev.left, right=side, top=prev.top, bottom=prev.bottom
            )
