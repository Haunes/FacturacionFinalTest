from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

class ExcelStyleManager:
    """Maneja los estilos para archivos Excel."""
    
    def __init__(self):
        # Fuentes
        self.header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        self.total_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        self.data_font = Font(name='Calibri', size=11, bold=False, color="000000")
        self.title_font = Font(name='Calibri', size=12, bold=True, color="000000")
        self.info_font = Font(name='Calibri', size=11)
        self.italic_font = Font(name='Calibri', size=9, italic=True, color="000000")
        
        # Colores
        HEADER_HEX = "FF002060"  # Azul corporativo
        self.header_fill = PatternFill(fill_type="solid", start_color=HEADER_HEX, end_color=HEADER_HEX)
        self.total_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        self.white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Bordes
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Alineaciones
        self.center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left = Alignment(horizontal='left', vertical='center', wrap_text=False)
        self.right = Alignment(horizontal='right', vertical='center', wrap_text=False)
    
    def style_cell(self, cell, text, font, alignment, fill=None, border=None):
        """Aplica estilos completos a una celda."""
        cell.value = text
        cell.font = font
        cell.alignment = alignment
        if fill is not None:
            cell.fill = fill
        if border is not None:
            cell.border = border
