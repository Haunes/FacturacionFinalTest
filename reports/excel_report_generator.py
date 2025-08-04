import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime
from utils.data_utils import get_document_count, find_column
from .excel_styles import ExcelStyleManager
from .excel_sheet_builder import ExcelSheetBuilder

class ExcelReportGenerator:
    """Generador de reportes en formato Excel para Ravago."""
    
    def __init__(self):
        self.style_manager = ExcelStyleManager()
        self.sheet_builder = ExcelSheetBuilder()
    
    def create_ravago_report(self, data: pd.DataFrame, anio: int, mes: str, funcionarios: dict = None) -> BytesIO:
        """
        Genera un Excel con dos hojas para Ravago.
        
        Args:
            data: Datos filtrados
            anio: Año del reporte
            mes: Mes del reporte
            funcionarios: Información de funcionarios
            
        Returns:
            Buffer con el archivo Excel generado
        """
        wb = Workbook()
        
        # Preparar datos auxiliares
        report_data = self._prepare_report_data(data, anio, mes, funcionarios)
        
        # Crear hoja de Facturación
        ws1 = wb.active
        ws1.title = "Facturación"
        self.sheet_builder.build_facturacion_sheet(ws1, report_data)
        
        # Crear hoja de Anexo 1
        ws2 = wb.create_sheet(title="Anexo 1")
        self.sheet_builder.build_anexo_sheet(ws2, report_data, data)
        
        # Guardar en memoria
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _prepare_report_data(self, data: pd.DataFrame, anio: int, mes: str, funcionarios: dict = None) -> dict:
        """Prepara los datos auxiliares para el reporte."""
        funcionarios = funcionarios or {}
        
        try:
            num_docs = get_document_count(data)
        except Exception:
            num_docs = len(data)
        
        # Calcular valor total
        if 'VALOR' in data.columns:
            total_valor = float(data['VALOR'].fillna(0).sum())
            valor_col = 'VALOR'
        else:
            valor_col = find_column(data, ['VALOR', 'TOTAL', 'IMPORTE', 'MONTO'])
            total_valor = float(data.get(valor_col, pd.Series(dtype=float)).fillna(0).sum()) if valor_col else 0.0
        
        # Encontrar columnas relevantes
        nombre_col = find_column(data, ['NOMBRE', 'NOMBRE CONTRAPARTE', 'CLIENTE'])
        tipo_doc_col = find_column(data, ['TIPO DE DOCUMENTO', 'TIPO DOCUMENTO', 'DOCUMENTO'])
        
        return {
            'anio': anio,
            'mes': mes,
            'num_docs': num_docs,
            'total_valor': total_valor,
            'valor_col': valor_col,
            'nombre_col': nombre_col,
            'tipo_doc_col': tipo_doc_col,
            'rep_name': funcionarios.get("reporta", "________________"),
            'rev_name': funcionarios.get("revisor", "________________"),
            'fecha_dt': funcionarios.get("fecha", datetime.now())
        }
